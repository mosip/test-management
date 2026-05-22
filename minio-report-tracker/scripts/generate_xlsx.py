import os
import json
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


def load_normalized_data(csv_path):
    rows = []
    with open(csv_path, 'r') as f:
        for line in f:
            line = line.strip()
            if not line or line.startswith("###"):
                continue
            parts = [p.strip() for p in line.split(',') if p.strip()]
            for i in range(0, len(parts), 8):
                if i + 7 < len(parts):
                    rows.append(parts[i:i + 8])

    df = pd.DataFrame(rows, columns=["Date", "Module", "T", "P", "S", "F", "I", "KI"])
    df = df[df["Date"].str.contains(r'\d{1,2}-[A-Za-z]+-\d{4}', na=False)]
    df["Date"] = pd.to_datetime(df["Date"], format="%d-%B-%Y")
    for col in ["T", "P", "S", "F", "I", "KI"]:
        df[col] = pd.to_numeric(df[col], errors='coerce')
    return df


def generate_graphs(df, output_dir):
    graph_files = []
    for module in df["Module"].unique():
        module_df = df[df["Module"] == module]
        if module_df.empty:
            continue

        module_df = module_df.groupby("Date").agg({"T": "sum", "P": "sum", "F": "sum"}).reset_index()
        module_df = module_df[module_df["Date"].dt.weekday < 5].sort_values("Date")
        if module_df.empty:
            continue

        plt.figure(figsize=(10, 5))
        plt.plot(module_df["Date"], module_df["T"], label="Total",  linewidth=2,   color='blue',  marker='o')
        plt.plot(module_df["Date"], module_df["P"], label="Passed", linewidth=2.5, color='green', marker='o')
        plt.plot(module_df["Date"], module_df["F"], label="Failed", linewidth=2.5, color='red',   marker='o')
        plt.title(f"Trend for Module: {module}", fontsize=14)
        plt.xlabel("Date", fontsize=12)
        plt.ylabel("Count", fontsize=12)
        plt.grid(True, linestyle='--', alpha=0.5)
        plt.legend()
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%d-%b-%Y'))
        plt.gca().xaxis.set_major_locator(mdates.DayLocator())
        plt.tight_layout()

        file_path = os.path.join(output_dir, f"{module}_trend.png")
        plt.savefig(file_path)
        graph_files.append((module, file_path))
        plt.close()

    return graph_files


def export_to_excel(csv_path, graph_files, xlsx_path, failure_reason=""):
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Module Data"

    df_raw = pd.read_csv(csv_path, dtype=str)
    df_raw = df_raw.loc[:, ~df_raw.columns.str.contains('^Unnamed', na=False)]
    df_raw = df_raw.dropna(axis=1, how='all')

    def normalize_header(col):
        return col.split('.')[0].strip()
    df_raw.columns = [normalize_header(col) for col in df_raw.columns]

    block_size = 8
    blocks = [df_raw.iloc[:, i:i + block_size] for i in range(0, df_raw.shape[1], block_size)]

    start_col = 1
    for block in blocks:
        for col_idx, col_name in enumerate(block.columns):
            cell = ws_data.cell(row=1, column=start_col + col_idx)
            cell.value = col_name
            cell.font = Font(bold=True)

        # Add failure reason as hover comment on the Date header cell of the first block
        if failure_reason and start_col == 1:
            c = Comment(failure_reason, "Workflow")
            c.width  = 300
            c.height = max(60, 40 + failure_reason.count('\n') * 18)
            ws_data.cell(row=1, column=start_col).comment = c

        for row_idx, row in enumerate(block.itertuples(index=False), start=2):
            for col_idx, value in enumerate(row):
                ws_data.cell(row=row_idx, column=start_col + col_idx, value=value)

        start_col += block_size + 1

    for col_cells in ws_data.columns:
        max_len = max(len(str(cell.value)) if cell.value else 0 for cell in col_cells)
        ws_data.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 2

    ws_charts = wb.create_sheet(title="Module Graphs")
    row_pos = 1
    for module, image_path in graph_files:
        if not os.path.exists(image_path):
            continue
        img = Image(image_path)
        img.width  = 800
        img.height = 400
        ws_charts.add_image(img, f"A{row_pos}")
        row_pos += 22

    wb.save(xlsx_path)


def create_failure_excel(alias, reason, xlsx_path):
    """Creates a minimal Excel for an env that failed entirely (no CSV produced)."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Status"

    red_fill = PatternFill("solid", start_color="FFCCCC")

    ws["A1"] = "Environment"
    ws["B1"] = alias
    ws["A1"].font = Font(bold=True)

    ws["A2"] = "Status"
    ws["B2"] = "FAILED"
    ws["B2"].font = Font(bold=True, color="FF0000")
    ws["B2"].fill = red_fill

    ws["A3"] = "Reason"
    ws["A3"].font = Font(bold=True)
    ws["B3"] = reason

    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 70

    wb.save(xlsx_path)


# === MAIN ===
csv_dir     = "minio-report-tracker/csv"
status_dir  = "status"
output_base = "minio-report-tracker/xlxs"
os.makedirs(output_base, exist_ok=True)

# Build reason map from status JSONs
reasons = {}
if os.path.isdir(status_dir):
    for sf in os.listdir(status_dir):
        if not sf.endswith(".json"):
            continue
        alias_name = sf.replace("status_", "").replace(".json", "")
        try:
            with open(os.path.join(status_dir, sf)) as f:
                data = json.load(f)
            reason = data.get("reason", "").strip()
            if reason:
                reasons[alias_name] = reason
        except (json.JSONDecodeError, IOError):
            pass

# Process envs that have a CSV
processed = set()
for file in os.listdir(csv_dir):
    if not file.endswith(".csv"):
        continue

    alias    = file.replace(".csv", "")
    csv_path = os.path.join(csv_dir, file)
    processed.add(alias)

    df_normalized  = load_normalized_data(csv_path)
    output_dir     = os.path.join(output_base, f"{alias}_images")
    os.makedirs(output_dir, exist_ok=True)

    graph_files    = generate_graphs(df_normalized, output_dir)
    xlsx_path      = os.path.join(output_base, f"{alias}.xlsx")
    failure_reason = reasons.get(alias, "")

    export_to_excel(csv_path, graph_files, xlsx_path, failure_reason=failure_reason)

# Create placeholder Excel for envs that failed entirely (no CSV produced)
for alias_name, reason in reasons.items():
    if alias_name not in processed:
        xlsx_path = os.path.join(output_base, f"{alias_name}.xlsx")
        create_failure_excel(alias_name, reason, xlsx_path)
